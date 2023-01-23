import re
import openpyxl as xl
import pandas as pd

class Formating(object):
    
    def __init__(self, sheet):
        raw_name = str(sheet)
        split_x = re.split(" ", raw_name, 1) 
        split_x2 = re.sub(">", "", split_x[1])
        self.sheet_name = split_x2.replace('"', "")
        
    def titles_n_codes(self, sheet):
        titles = []
        codes = []
        if self.sheet_name == 'Property and Transaction':
            for i in sheet.iter_rows(min_row = 5, max_col = 33, max_row = 5, values_only = True):
                titles.append(i)
                
            for i in sheet.iter_rows(min_row = 7, max_col = 33, max_row = 7, values_only = True):
                codes.append(i)
                
        if self.sheet_name == "Valuation and Financial":
            for i in sheet.iter_rows(min_row=5, max_col = 62, max_row = 5, values_only = True):
                titles.append(i)
                
            for i in sheet.iter_rows(min_row=7, max_col = 62, max_row = 7, values_only = True):
                codes.append(i)
                
        titles = list(map(str.lower, titles[0][:]))
        codes = list(map(str.lower, codes[0][:]))
        
        
    def columndiff(og_titles, t_titles):
        title_check = []
        for i in og_titles[0][:]:
            i = str(i)
            for x in t_titles[0][:]:
                x = str(x)
                if i == x:
                    title_check.append(i)
        
        s = set(title_check)
        different_cols = [x for x in og_titles[0][:] if x not in s]
        return different_cols
    
    def columncheck2(og_titles, og_code):
        df_all = (pd.DataFrame([og_titles, og_code])).T
        counter = []
        for i in range(len(df_all)):
            counter.append(i)
        df_all['Position'] = counter
        df_all = (df_all.rename(columns={0:"Columns", 1:"Code"})).set_index("Columns")
        df_all['Code'] = df_all['Code'].str.lower()
        df_all = df_all.T
        df_all = df_all.rename(index={1:"Code"}) 
        return df_all
    
    def columncheck(og_titles, og_code):
        df_all = (pd.DataFrame([og_titles[0], og_code[0]])).T
        counter = []
        for i in range(len(df_all)):
            counter.append(i)
        df_all['Position'] = counter
        df_all = (df_all.rename(columns={0:"Columns"})).set_index("Columns")
        df_all = df_all.T
        df_all = df_all.rename(index={1:"Code"})
        return df_all
    
    def copypasteregular2(og_sheet, t_sheet, og_df_all, t_df_all):
        og_positions = []
        t_positions = []
        for i in og_df_all.columns:
            for x in t_df_all.columns:
                if i == x:
                    og_positions.append(og_df_all[i])
                    t_positions.append(t_df_all[x])
        og_positions = pd.DataFrame(og_positions)
        t_positions = pd.DataFrame(t_positions)
        pairs = list(zip(og_positions['Position'], t_positions['Positions']))
        print(pairs) 
        for z, y in pairs:
            range_selected = Formating.copyRange(z, 8, z, 300, og_sheet)
            Formating.pasteRange(y, 8, y, 300, t_sheet, range_selected)
            
    def copypasteirregular(og_sheet, t_sheet, og_df_all, t_df_all):
        missing_cols = [x for x in og_df_all.columns if x not in t_df_all.columns]
        column_name_miss = []
        column_pos = []
        column_code = []
        for i in range(len(og_df_all.loc['Code'])):
            column_name_miss.append(og_df_all.columns[i])
            column_pos.append(og_df_all.loc["Position"][i])
            column_code.append(og_df_all.loc["Code"][i])
        sub_data = (pd.DataFrame([column_name_miss, column_pos, column_code])).T
        sub_data = sub_data.rename(columns={0: "Column Name", 1: "Position", 2 : "Code"})
        
        intersect = set(list(og_df_all.loc["Code"])).intersection(set(list(sub_data["Code"])))
        intersect = list(intersect)
        
        t_df_all_trans = t_df_all.T
        for x in range(len(intersect)):
            for i in t_df_all_trans["Position"].loc[t_df_all_trans["Code"] == intersect[x]]:
                val = int(i)
                
                if val == 0:
                    val += 1
                    
                range_selected_ir = Formating.copyrange(val, 8, val, 100, og_sheet)
                Formating.pasterange(val, 8, val, 100, t_sheet, range_selected_ir)
        return sub_data 
    
    
    def copyrange(startcol, startrow, endcol, endrow, sheet):
        rangeselected = []
        # Loop Through Rows 
        for i in range(startrow, endrow + 1, 1):
            rowselected = []
            # Appends Row to a Row selected List
            for j in range(startcol, endcol + 1, 1):
                rowselected.append(sheet.cell(row=i, column=j).value)
            # Adds the row selected list and nests inside the range selected
            rangeselected.append(rowselected)
        return rangeselected
                                                      
    def pasterange(startcol, startrow, endcol, endrow, sheetreceiving, copieddata):
        countrow = 0
        for i in range(startrow, endrow + 1, 1):
            countcol = 0
            for j in range(startcol, endcol + 1, 1)
            sheetreceiving.cell(row=i, column=j).value = copieddata[countrow][countcol]
            countcol += 1
        countrow += 1
        
    def copypasteregular3(og_sheet, t_sheet, og_df_all, t_df_all):
        t_df_all_trans = t_df_all.T.reset_index().set_index("Code")
        og_df_all_trans = og_df_all.T.reset_index().set_index("Code")
        
        merge = t_df_all_trans.merge(og_df_all_trans, left_index = True, right_index = True,
                                     suffixes=('_Template', '_Original'))
        
        for index, row in merge[1:].iterrows():
            range_selected = Formating.copyRange(row['Position_Original'], 8, row['Position_Original'], 300, og_sheet)
            Formating.pasterange(row['Position_Template'], 8, row['Position_Template'], 300, t_sheet, range_selected)
        return merge
    
    def titles_n_codes2(self, sheet):
        titles = []
        codes = []
        titles_final = []
        codes_final =[]
        if self.sheet_name == "Property and Transaction":
            for i in sheet.iter_rows(min_row=5, max_col = 31, max_row= 5, values_only =True):
                titles.append(i)
            
            for i in sheet.iter_rows(min_row=7, max_col = 31, max_row = 7, values_only = True):
                codes.append(i)
            
            for i in codes[0][:]:
                i = str(i)
                codes_final.append(i)
            
            for i in titles[0][:]:
                i = str(i)
                titles_final.append(i)
                
        return titles_final, codes_final
        
    if self.sheet_name == "Valuation and Financial":
        for i in sheet.iter_rows(min_row=5, max_col = 62, max_row= 5, values_only=True):
            titles.append(i)
        
        for i in sheet.iter_rows(min_row=7, max_col =62, max_row =7, values_only = True):
            codes.append(i)
        
        for i in codes[0][:]:
            i = str(i)
            codes_final.append(i)
            
        for i in titles[0][:]:
            i = str(i)
            titles_final.append(i)
            
        my_dict_code = {key: 0 for key in codes_final}
        for i in range(len(my_dict_code)):
            my_dict_code[codes_final[i]] += 1
            if my_dict_code[codes_final[i]] > 1: 
                codes_final[i] = codes_final[i] + "_1" + str(my_dict_code[codes_final[i]] -  1)
            
        return titles_final, codes_final
    
    def copypasteregular3(og_sheet, t_sheet, og_df_all, t_df_all):
        t_df_all_trans = t_df_all.T.reset_index().set_index("Code")
        og_df_all_trans = og_df_all.T.reset_index().set_index("Code")
        merge_df = t_df_all_trans.merge(og_df_all_trans, left_index = True, right_index = True,
                                        suffixes = ('_Template', '_Original'))
        for x, y in zip(merge_df['Position_Original'], merge_df['Position_Template']):
            if x or y != 0:
                range_selected = Formating.copyrange(x, 8m x, 300, og_sheet)
                Formating.pasterange(y, 8, y, 300, t_sheet, range_selected)
        return merge_df
    
    
    def fund_recon(recon_sheet):
        recon_sheet['D12'] = "=ABS(SUM(L42:L342)) * -1" # Negative sum of debt market value
        recon_sheet['D13'] = "=-1 * F29" # Negative debt market value fund level data F29
        recon_sheet['D11'] = "=SUM(D42:D342)" # Sum of market value
        recon_sheet['D14'] = '=M29' # Cash Balance fund level data
        recon_sheet['D15'] = "=029" # Other assets fund level data
        recon_sheet['D16'] = "=ABS(P29) * -1" # Other liabilities negative
        recon_sheet['D18'] = "=SUM(D11:D16)" # Implied Net Asset Value
        recon_sheet['D19'] = "=D33" # Stated NAV
        recon_sheet['D20'] = "=D19- D18"
        recon_sheet['D21'] = "=D20/D19"
        
        # Income
        recon_sheet['G11'] = "=SUM(J42:J342)"
        recon_sheet['G12'] = "ABS(SUM(N42:N342))"
        recon_sheet['G13'] = "=-ABS(K29)"
        recon_sheet['G14'] = "=ABS(Q29) * -1
        recon_sheet['G15'] = "=R29"
        recon_sheet['G16'] = "=E33 - H33"
        recon_sheet['G18'] = "=SUM(G11:G16)"
        recon_sheet['G19'] = "=E33"
        recon_sheet['G20'] = "=-(G19 - G18)"
        recon_sheet['G21'] = "=G20/G19"
        
        # Appreciation
        recon_sheet['J11'] = "=SUM(F42 : F342)"
        recon_sheet['J12'] = "=SUM(M42: M342)"
        recon_sheet['J13'] = "=G29"
        recon_sheet['J14'] = "=F33 - I33"
        recon_sheet['J18'] = "=SUM(J11:J14)"
        recon_sheet['J19'] = "=F33"
        recon_sheet['J20'] = "=(J19 - J18) * 1 " 
        recon_sheet['J21'] = "=J20/J19"
        
        
original_file = input()
original = load_workbook(original_file, data_only = True)
og_pt_sheet= original["Property and Transaction"]
template = load_workbook("US Quarterly Data Template 2021.xlsx", data_only = True)
t_pt_sheet = template["Property and Transaction"]

original_property_transaction = Formating(og_pt_sheet)
og_titles_p, og_code_p = Formating.titles_n_codes2(original_property_transaction, og_pt_sheet)
og_df_all_p = Formating.columncheck2(og_titles_p, og_code_p)

template_pt = Formating(t_pt_sheet)
t_titles_p, t_code_p = Formating.titles_n_codes2(template_pt, t_pt_sheet)
t_df_all_p = Formating.columncheck2(t_titles_p, t_code_p)
merge_p = Formating.copypasteregular3(og_pt_sheet, t_pt_sheet, og_df_all_p, t_df_all_p)

og_f_sheet = original["Valuation and Financial"]
t_f_sheet = template["Valuation and Financial"]

original_financial = Formating(og_f_sheet)
og_titles_f, og_code_f = Formating.titles_n_codes2(original_financial, og_f_sheet)
og_df_all_f = Formating.columncheck2(og_titles_f, og_code_f)

template_f = Formating(t_f_sheet)
t_titles_f, t_code_f = Formating.titles_n_codes2(template_f, t_f_sheet)
t_df_all_f = Formating.columncheck2(t_titles_f, t_code_f)
merge_f = Formating.copypasteregular3(og_f_sheet, t_f_sheet, og_df_all_f, t_df_all_f)

cap_rate_range = Formating.copyrange(20,8,20,300, t_f_sheet) 
Formating.pasterange(18,8,18,300, t_f_sheet, cap_rate_range)
delete_rent = Formating.copyrange(100,8,100,300, t_f_sheet)
Formating.pasterange(20, 8, 20, 300, t_f_sheet, delete_rent)
let_area = Formating.copyrange(46, 8, 46, 300, t_f_sheet)
number_of_units = Formating.copyrange(47, 8, 47, 300, t_f_sheet)

Formating.pasterange(45, 8, 45, 300, t_f_sheet, let_area)
Formating.pasterange(46, 8, 46, 300, t_f_sheet, number_of_units)
Formating.pasterange(47, 8, 47, 300, t_f_sheet, percentage_rented)

og_fld_sheet = original["Fund Level Data"]
t_fld_sheet = template["Fund Level Data"]
range_selected_fld = Formating.copyRange(1, 8, 26, 30, og_fld_sheet)
Formating.pasterange(1, 8, 26, 15, t_fld_sheet, range_selected_fld)

og_twr_sheet = original["Fund TWR"]
t_twr_sheet = template["Fund TWR"]
range_selected_twr = Formating.copyrange(1, 7, 30, 100, og_twr_sheet)
Formating.pasteRange(1, 7, 30, 100, t_twr_sheet, range_selected_twr)

t_recon_sheet = template["Fund Recon"]
Formating.fund_recon(t_recon_sheet)

cap_rate_range = Formating.copyRange(20,8,20, 300, t_f_sheet)
Formating.pasterange(18,8,18,300, t_f_sheet, cap_rate_range)
delete_rent = Formating.copyrange(100, 8, 100, 300, t_f_sheet)
Formating.pasterange(20, 8, 20, 300, t_f_sheet, delete_rent)
let_area = Formating.copyrange(46, 8, 46, 300, t_f_sheet)
percentage_rented = Formating.copyrange(45, 8, 45, 300, t_f_sheet)
number_of_units = Formating.copyRange(47, 8, 47, 300, t_f_sheet)

Formating.pasterange(45, 8, 45, 300, t_f_sheet, let_area)
Formating.pasterange(46, 8, 46, 300, t_f_sheet, number_of_units)
Formating.pasterange(47, 8, 47, 300, t_f_sheet, percentage_rented)

fund_twr = Formating.copyrange(2, 7, 19, 7, t_twr_sheet)
Formating.pasterange(1, 33, 18, 33, t_recon_sheet, fund_twr)
t_recon_sheet['J35'] = "=E33 - H33"
t_recon_sheet['J36'] = "=F33 - I33"
t_recon_sheet['J37'] = "=J35 + J36"

fld_ol = Formating.copyrange(2, 8, 17, 8, t_fld_sheet)
Formating.pasterange(1, 29, 16, 29, t_recon_sheet, fld_ol)
inc_state = Formating.copyrange(25, 8, 26, 8, t_fld_sheet)
Formating.pasterange(17, 29, 18, 29, t_recon_sheet, inc_state)

pld_2 = Formating.copyrange(2, 8, 3, 300, t_pt_sheet)
Formating.pasterange(2, 42, 3, 300, t_recon_sheet, pld_2)
market_val = Formating.copyrange(12, 8, 12, 300, t_f_sheet)
Formating.pasterange(4, 42, 4, 300, t_recon_sheet, market_val)
ps_pp = Formating.copyrange(21, 8, 22, 300, t_f_sheet)
Formating.pasterange(7, 42, 8, 42, t_recon_sheet, ps_pp)
cap_improv = Formating.copyrange(24, 8, 24, 300, t_f_sheet)
Formating.pasterange(9,42,9,300, t_recon_sheet, cap_improv)
debt_bv = Formating.copyrange(48, 8, 50, 300, t_f_sheet)
Formating.pasterange(11, 42, 13, 300, t_recon_sheet, debt_bv)
net_operating_inc = Formating.copyrange(30, 8, 30, 300, t_f_sheet)
Formating.pasterange(10, 42, 10, 300, t_recon_sheet, net_operating_inc)
int_on_debt = Formating.copyrange(53, 8, 53, 300, t_f_sheet) 
Formating.pasterange(14, 42, 14, 300, t_recon_sheet, int_on_debt) 

template.save(r"C:\Users\pereeug\OneDrive\OneDrive - MSCI Office 365\Documents\Real Estate\Automate_analysis\Templates\Output\Working_file.xlsx")
                     