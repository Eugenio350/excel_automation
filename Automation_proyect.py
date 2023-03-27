#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import json
import regex as re
import openpyxl as xl
import pandas as pd
import ipywidgets as widgets
from IPython.display import display
import io
from ipywidgets import interact, widgets, Layout, interactive

f = open(r'/home/jovyan/work/shared_win_otw/voila/Eugenio/Real_estate/Resources/config_ldct.json')
config = json.load()
config_p = config["Property And Transaction"]
config_v = config["Valuation and Financial"]
config_f = config["Fund Level Data"]
config_general = config["General"]
config_merge = config["Merge Config"]
config_recon = config["Recon"]

def display_data():
    w = widgets.IntSlider()
    uploader = widgets.FileUpload()
    specification = widgets.Dropdown(options = ['Regular File', 'Irregular File'],
                                    description = 'Type of File',
                                    disabled = False)
    file_name = widgets.Textarea(value = "Text Name", description = "File name", disabled = False)
    saved_file_name = widgets.Textarea(value = "Text Name", description = "Processed file name", disabled = False)
    select_button = widgets.Button(description = 'Search', disabled = False, button_style = 'primary',
                                  layout = Layout(margin=('0px 0px 0px 10px'), width = '10%'))
    
    h_box = widgets.HBox([uploader, specification, file_name, select_button, saved_file_name])
    display(h_box)
    
    class Formating(object):
        
        def __init__(self, sheet):
            raw_name = str(sheet)
            split_x = re.split(" ", raw_name, 1)
            split_x2 = re.sub(">","",split_x[ 1 ])
            self.sheet_name = split_x2.replace('"', "")
            
        def Titles_n_Codes2(self, sheet):
            titles = []
            codes = []
            titles_final = []
            codes_final = []
            
            if self.sheet_name == 'Property and Transaction':
                for i in sheet.iter_rows(min_row = config_p['t_min_row'], max_col = config_p['t_max_col'], max_row = config_p['t_max_row'], values_only = True):
                    titles.append(i)
                
                for i in sheet.iter_rows(min_row = config_p['c_min_row'], max_col = config_p['c_max_col'],
                                        max_row = config_p['c_max_row'], values_only = True):
                    codes.append(i)
                    
                for i in codes[0][:]:
                    i = str(i)
                    codes_final.append(i)
                    
                for i in titles[0][:]:
                    i = str(i)
                    titles_final.append(i)
                    
                return titles_final, codes_final
            
            if self.sheet_name == "Valuation and Financial":
                for i in sheet.iter_rows(min_row=config_v['t_min_row'], max_col=config_v['t_max_col'],
                                        max_row = config_v['t_max_row'], values_only = True):
                    titles.append(i)
                    
                for i in sheet.iter_rows(min_row = config_v['c_min_row'], max_col = config_v['c_max_col'],
                                        max_row= config_v['c_max_row'], values_only=True):
                    codes.append(i)
                
                for i in codes[ 0 ][ : ]:
                    i = str(i)
                    codes_final.append(i)
                
                for i in titles[0][:]:
                    i = str(i)
                    titles_final.append(i)
                    
            if self.sheet_name == "Fund Level Data":
                for i in sheet.iter_rows(min_row=config_f['t_min_row'], max_col=config_f['t_max_col'],
                                        max_row = config_f['t_max_row'], values_only = True):
                    titles.append(i)
                
                for i in sheet.iter_rows(min_row=config_f['c_min_row'], max_col=config_f['c_max_col'],
                                        max_row = config_f['c_max_row'], values_only = True):
                    codes.append(i)
                    
                for i in codes[0][:]:
                    i = str(i)
                    codes_final.append(i)
                    
                for i in titles[0][:]:
                    i = str(i)
                    titles_final.append(i)
                    
            my_dict_code = {key: 0 for key in codes_final}
            for i in range(len(my_dict_code)):
                my_dict_code[codes_final[i]] += 1
                if my_dict_code[ codes_final[ i ]] > 1:
                    codes_final[ i ] = codes_final[ i ] + "_1" + str(my_dict_code[ codes_final[ i] ] -1)
                    
            return titles_final, codes_final
        
        def ColumnCheck2(og_titles, og_code):
            df_all = (pd.DataFrame([og_titles, og_code])).T
            counter = []
            for i in range(len(df_all)):
                counter.append(i + 1)
            df_all['Position'] = counter
            df_all = (df_all.rename(columns={0: "Columns", 1:"Code"})).set_index("Columns")
            df_all['Code'] = df_all['Code'].str.lower()
            df_all = df_all.T
            df_all = df_all.rename(index={1:"Code"})
            return df_all
        
        def copyRange(startCol, startRow, endCol, endRow, sheet):
            rangeSelected = [ ]
            #Loop Through Rows
            for i in range(startRow, endRow + 1, 1):
                # Appends Row to a Row Selected List
                rowSelected = [ ]
                for j in range(startCol, endCol + 1, 1):
                    rowSelected.append(sheet.cell(row=i, column=j).value)
                # Adds the RowSelected list and nests inside the range selected
                rangeSelected.append(rowSelected)
            return rangeSelected
                        
        def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
            countRow = 0
            for i in range(startRow, endRow + 1, 1):
                countCol = 0
                for j in range(startCol, endCol + 1, 1):
                    sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
                    countCol += 1
                countRow += 1
                
        def CopyPasteRegular3(og_sheet, t_sheet, og_df_all, t_df_all):
            t_df_all_trans = t_df_all.T.reset_index().set_index("Code")
            og_df_all_trans = og_df_all.T.reset_index().set_index("Code")
            non_present_code = [x for x in t_df_all_trans.index if x not in og_df_all_trans.index]
            
            merge_df = t_df_all_trans.merge(og_df_all_trans, left_index = True, right_index = True, 
                                           suffixes=('_Template', '_Original'))
            
            for x, y in zip(merge_df['Position Original'], merge_df['Position_Template']):
                if x or y != 0:
                    range_selected = Formating.copyRane(x, 8, x, config_general['all_row_number'], og_sheet)
                    
                    Formating.pasteRange(y, 8, y, config_general['all_row_number'], t_sheet, range_selected)
                    
            return merge_df
        
        def fund_recon(recon_sheet):
            recon_sheet['D12'] = "=ABS(SUM(L42:L342)) * -1"
            recon_sheet['D13'] = "=-1 * F29"
            recon_sheet['D14'] = "=M29"
            recon_sheet['D15'] = "=O29"
            recon_sheet['D16'] = "=ABS(P29) * -1"
            recon_sheet['D18'] = "=SUM(D11:D16)"
            recon_sheet['D19'] = "=D33"
            recon_sheet['D20'] = "=D19 - D18"
            recon_sheet["D21"] = "=D20/D19"
            
            #Income 
            recon_sheet['G11'] = "=SUM(J42:J342)"
            recon_sheet['G12'] = "=ABS(SUM(N42:N342)) * -1"
            recon_sheet['G13'] = "=-ABS(K29)"
            recon_sheet['G14'] = "=ABS(Q29) * -1"
            recon_sheet['G15'] = "=R29"
            recon_sheet['G16'] = "=E33 - H33"
            
        def twr_paste_condition(og_sheet_twr, t_sheet):
            #Reorganize Dates
            twr_dates = []
            for i in og_sheet_twr.iter_rows(min_row = config_twr['c_dates_min_row'], min_col = config_twr['c_dates_min_col'],
                                           max_col = config_twr['c_dates_max_col'], max_row = config_general['all_row_number'], values_only =True): 
                twr_dates.append(i)
            twr_dates = list(map(list, twr_dates))
            for i in reversed(twr_dates):
                if i == [None]:
                    twr_dates.pop()
                if i != [None]:
                    break
        
        def last_q_paste(template, df_file, lastq_file):
            last_q = xl.load_workbook(lastq_file, data_only = True)
            last_sheet_names = last_q.sheetnames
            last_q_df = pd.read_excel(lastq_file, sheet_name = last_sheet_names[1])
            last_q_df[["Client Asset ID", "Capital Value"]]
            id_list = df_file["Client Property Reference #"]
            last_q_df_clean = last_q_df.loc[last_q_df["Client Asset ID"].isin(id_list)]
            last_q_df_clean = last_q_df_clean.rename(columns={"Client Asset ID": :"Client Property Reference #",
                                                             "Capital Value": "Previous Market Value"})
            merge_result = df_file.merge(last_q_df_clean, right_on = "Client Property Reference #", left_on = "Client Property Reference #",
                                        how = 'inner')
            row_len = len(merge_result)
            return merge_result
        
    def investigation(a):
        try:
            original_file = str(Ldct_name.value)
            print(original_file)
            
            original = xl.load_workbook(original_file, data_only = True)
            template = xl.load_workbook("US Quarterly Data Template 2021.xlsx", data_only = True)
            og_pt_sheet = original["Property and Transaction"]
            t_pt_sheet = template["Property and Transaction"]
            
            original_property_transaction = Formating(og_pt_sheet)
            og_titles_p, og_code_p = Formating.Titles_n_Codes2(original_property_transaction, og_pt_sheet)
            og_df_all_p = Formating.ColumnCheck2(og_titles_p, og_code_p)
            
            template_pt = Formating(t_pt_sheet)
            t_titles_p, t_code_p = Formating.Titles_n_Codes2(template_pt, t_pt_sheet)
            t_df_all_p = Formating.ColumnCheck2(t_titles_p, t_code_p)
            merge_p = Formating.CopyPasteRegular3(og_pt_sheet, t_pt_sheet, og_df_all_p, t_df_all_p)
            
            og_f_sheet = original["Valuation and Financial"]
            t_f_sheet = template["Valuation and Financial"]
            
            t_template_financial = Formating(t_f_sheet)
            t_titles_f, t_code_f = Formating.Titles_n_Codes2(t_template_financial, t_f_sheet)
            t_df_all_f = Formating.ColumnCheck2(t_titles_f, t_code_f)
            
            original_financial = Formating(og_f_sheet)
            og_titles_f, og_code_f = Formating.Titles_n_Codes2(original_financial, og_f_sheet)
            og_df_all_f = Formating.ColumnCheck2(og_titles_f, og_code_f)
            merge_f = Formating.CopyPasteRegular3(og_f_sheet, t_f_sheet, og_df_all_f, t_df_all_f)
            
            og_fld_sheet = original["Fund Level Data"]
            t_fld_sheet = template["Fund Level Data"]
            
            original_fld = Formating(og_fld_sheet)
            og_titles_fld, og_code_fld = Formating.Titles_n_Codes2(original_fld, og_fld_sheet)
            og_df_all_fld = Formating.ColumnCheck2(og_titles_fld, og_code_fld)
            
            t_template_fld = Formating(t_fld_sheet)
            t_titles_fld, t_code_fld = Formating.Titles_n_Codes2(t_template_fld, t_fld_sheet)
            t_df_all_fld = Formating.ColumnCheck2(t_titles_fld, t_code_fld)
            merge_fld = Formating.CopyPasteRegular3(og_fld_sheet, t_fld_sheet, og_df_all_fld, t_df_all_fld)
            
            t_recon_sheet = template["Fund Recon"]
            
            Formating.fund_recon(t_recon_sheet)
            
            fld_ol = Formating.copyRange(config_recon['c_fld_ol_min_col'], config_recon['c_fld_ol_min_row'],
                                        config_recon['c_fld_ol_max_col'], config_recon['c_fld_ol_max_row'], t_fld_sheet)
            
            Formating.pasteRange(1,29,16,29, t_recon_sheet, fld_ol)
            inc_state = Formating.copyRange(25, 8, 26, 8, t_fld_sheet)
            Formating.pasteRange(17,29,18,29, t_recon_sheet, inc_state)
            
            # TWR DATA AND PROPERTY LEVEL DATA
            
            # RANGES SHOULDNT BE HARDCODED WHEN POSSIBLE
            
            pld_2 = Formating.copyRange(2, 8, 3, config_general['all_row_number'], t_pt_sheet)
            Formating.pasteRange(2, 42, 3, config_general['all_row_number'], t_recon_sheet, pld_2)
            market_val = Formating.copyRange(12, 8, 12, config_general['all_row_number'], t_f_sheet)
            Formating.pasteRange(4, 42, 4, config_general['all_row_number'], t_recon_sheet, market_val)
            
            ps_pp = Formating.copyRange(21, 8, 22, config_general['all_row_number'], t_f_sheet)
            Formating.pasteRange(7, 42, 8, 42, t_recon_sheet, ps_pp)
            cap_improv = Formating.copyRange(24, 8, 24, config_general['all_row_number'], t_f_sheet)
            Formating.pasteRange(9, 42, 9, config_general['all_row_number'], t_recon_sheet, cap_improv)
            debt_bv = Formating.copyRange(48, 8, 50, config_general['all_row_number'], t_f_sheet)
            Formating.pasteRange(11, 42, 13, config_general['all_row_number'], t_recon_sheet, debt_bv)
            net_operating_inc = Formating.copyRange(30, 8, 30, config_general['all_row_number'], t_f_sheet)
            Formating.pasteRange(10, 42, 10, config_general['all_row_number'], t_recon_sheet, net_operating_inc)
            int_on_debt = Formating.copyRange(53, 8, 53, config_general['all_row_number'], t_f_sheet)
            Formating.pasteRange(14, 42, 14, config_general['all_row_number'], t_recon_sheet, int_on_debt)
            
            og_twr_sheet = original["Fund TWR"]
            t_twr_sheet = template["Fund TWR"]
            
            range_selected_twr = Formating.copyRange(1, 7, 19, 100, og_twr_sheet)
            Formating.pasteRange(1, 7, 19, 100, t_twr_sheet, range_selected_twr)
            twr_dates = Formating.twr_paste_condition(og_twr_sheet, t_recon_sheet)
            
            name_of_document = str(saved_ldct_name.value)
            
            data = t_f_sheet.values
            data_df = pd.DataFrame(data)
            data_df = data_df[4:].reset_index()
            del(data_df['index'])
            data_df.columns = data_df.iloc[0]
            data_df = data_df[2:].reset_index()
            del(data_df['index'])
            
            merge = Formating.last_q_paste(template, data_df, str(Last_q_name.value))
            row_len = len(merge)
            merge_reduced = merge[config_merge["Columns Recon"]]
            merge_reduced.to_excel("Merge_output.xlsx")
            time.sleep(5)
            
            merge_workbook = xl.load_workbook("Merge_output.xlsx ")
            merge_sheetnames = merge_workbook.sheetnames
            merge_sheet = merge_workbook[str(merge_sheetnames[0])]
            merge_paste = Formating.copyRange(2, 2, 5, int(row_len + 1), merge_sheet)
            Formating.pasteRange(2, 42, 5, 42 + int(row_len - 1), t_recon_sheet, merge_paste)
            template.save(r"/home/jovyan/work/shared_win_otw/voila/Eugenio/Real_estate" + "/" + name_of_document + '.xlsx')
            print('Done')
        except EOFError as e: 
            print(end = ' ')
    select_button.on_click(investigation)


# In[ ]:


def merge_two_rows(row1, ro2, static_cols, sum_cols, mean_cols):
    merged_dict = {}
    
    for col in row1.index:
        if col in static_cols:
            merged_dict[col] = row1[col]
            
        if col in sum_cols:
            merged_dict[col] = row1[col] + row2[col]
            
        elif col in mean_cols:
            merged_dict[col] = np.mean([row1[col], row2[col]])
        
        else:
            merged_dict[col] = row1[col]
        
    merged_df = pd.DataFrame(merged_dict, index=[0])
    return merged_df


# In[ ]:


import json
import re
import openpyxl as xl
import pandas as pd
import ipywidgets as widgets
from IPython.display import display
import io
from ipywidgets import interact, widgets, Layout, interactive

f = open('/home/jovyan/work/shared_win_otw/voila/Eugenio/Real_estate/Resources/config_ldct.json')
config = json.load(f)
config_p = config["Property And Transaction"]
config_v = config["Valuation and Financial"]
config_f = config["Fund Level Data"]
config_general = config["General"]
config_merge = config["Merge Config"]
config_recon = config["Recon"]

def display_data():
    w = widgets.IntSlider()
    uploader = widgets.FileUpload()
    specification = widgets.Dropdown(options = ['Regular File', 'Irregular File'],
                                    description = 'Type of File',
                                    disabled = False)
    file_name = widgets.Textarea(value = "Text Name", description = "File name", disabled = False)
    saved_file_name = widgets.Textarea(value = "Text Name", description = "Processed file name", disabled = False)
    select_button = widgets.Button(description = 'Search', disabled = False, button_style = 'primary',
                                  layout = Layout(margin=('0px 0px 0px 10px'), width = '10%'))
    
    h_box = widgets.HBox([uploader, specification, file_name, select_button, saved_file_name])
    display(h_box)
    
    class Formating(object):
        
        def __init__(self, sheet):
            raw_name = str(sheet)
            split_x = re.split(" ", raw_name, 1)
            split_x2 = re.sub(">","",split_x[ 1 ])
            self.sheet_name = split_x2.replace('"', "")
            
        def Titles_n_Codes2(self, sheet):
            titles = []
            codes = []
            titles_final = []
            codes_final = []
            
            if self.sheet_name == 'Property and Transaction':
                for i in sheet.iter_rows(min_row=config_p['t_min_row'], max_col=config_p['t_max_col'], max_row=config_p['t_max_row'], values_only=True):
                    titles.append(i)
                
                for i in sheet.iter_rows(min_row=config_p['c_min_row'], max_col=config_p['c_max_col'],
                                        max_row=config_p['c_max_row'], values_only=True):
                    codes.append(i)
                    
                for i in codes[0][:]:
                    i = str(i)
                    codes_final.append(i)
                    
                for i in titles[0][:]:
                    i = str(i)
                    titles_final.append(i)
                    
                return titles_final, codes_final
            
            if self.sheet_name == "Valuation and Financial":
                for i in sheet.iter_rows(min_row=config_v['t_min_row'], max_col=config_v['t_max_col'],
                                        max_row=config_v['t_max_row'], values_only=True):
                    titles.append(i)


# In[ ]:


def merge_rows(df, group_cols, agg_cols, static_cols, merge_func):
    """
    Merge rows in a pandas DataFrame based on a custom condition, while also aggregating specified columns and 
    keeping specified static columns based on the value of one row or another.
    
    Args:
        df (pandas.DataFrame): The input DataFrame to merge rows in.
        group_cols (list): A list of column names to group the DataFrame by.
        agg_cols (dict): A dictionary that specifies how to aggregate the columns to be merged, 
            where the keys are column names and the values are aggregation functions.
        static_cols (list): A list of column names to keep static based on the value of one row or another.
        merge_func (function): A function that takes a DataFrame group as input and returns a Boolean value indicating 
            whether the group should be merged.
    
    Returns:
        pandas.DataFrame: The merged DataFrame.
    """
    # Group the DataFrame by the specified columns and apply the merge function to each group to create a merge column
    df['merge'] = df.groupby(group_cols).apply(merge_func).reset_index(drop=True)
    
    # Filter out the rows that should not be merged
    filtered_df = df[~df['merge']].copy()
    
    # Group the remaining rows by the specified columns and aggregate the specified columns
    agg_df = filtered_df.groupby(group_cols).agg(agg_cols).reset_index()
    
    # Keep the static columns based on the value of the first row for each group of rows with the same group column values
    static_df = filtered_df.groupby(group_cols).first().reset_index()[group_cols + static_cols]
    
    # Merge the aggregated DataFrame and static DataFrame on the group columns
    merged_df = pd.merge(agg_df, static_df, on=group_cols)
    
    return merged_df

